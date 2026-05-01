"""
plant_report_updater.py
────────────────────────────────────────────────────────────────────────────
G-MES에서 다운로드한 LPA/5S raw data를
Plant_Mobile_Audit_Report.xlsx 의 LPA/5S 시트에 붙여넣고
Rates 시트의 SUMIFS 수식이 자동으로 재계산되도록 저장.

사용법:
  python plant_report_updater.py --lpa LPA파일.xls --5s 5S파일.xls
  python plant_report_updater.py  # lpa_5s_combined_sender.py와 연동 시 자동 실행

출력:
  output/Plant_Mobile_Audit_Report_YYYYMMDD.xlsx
────────────────────────────────────────────────────────────────────────────
"""
import sys
import os
import argparse
import shutil
import subprocess
import pandas as pd
import openpyxl

from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import load_workbook

# lpa_5s_combined_sender.env 로드
load_dotenv(Path(__file__).parent / 'lpa_5s_combined_sender.env')


# 프로젝트 루트 경로를 sys.path에 추가 (core import 전에 반드시 필요)
_PROJECT_ROOT = str(Path(__file__).resolve().parent.parent.parent)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

# === Core RPA 표준 함수 ===
from core.common_fn import (log, safe_filename)


# ══════════════════════════════════════════════════════════════════════
# ▌ 설정
# ══════════════════════════════════════════════════════════════════════

_project_root = Path(os.getenv("PROJECT_ROOT", str(Path(__file__).parent)))

# 템플릿 파일 경로 (원본 — 덮어쓰지 않음)
TEMPLATE_PATH = Path(os.getenv("TEMPLATE_PATH", str(_project_root / "31111 Plant_Mobile_Audit_Report.xlsx")))

# 출력 폴더
OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR_PRIMARY", str(_project_root / "output")))

# LibreOffice 경로 (수식 재계산용, 없으면 Excel에서 열 때 자동 계산)
LIBREOFFICE_PATH = os.getenv("LIBREOFFICE_PATH", r"C:\Program Files\LibreOffice\program\soffice.exe")


# ══════════════════════════════════════════════════════════════════════
# ▌ 1. raw data 로드
# ══════════════════════════════════════════════════════════════════════

def load_raw(filepath: Path, sheet_type: str) -> pd.DataFrame:
    """
    G-MES에서 다운로드한 LPA 또는 5S xls 파일 로드.
    헤더 행 자동 탐색.
    """
    filepath = Path(filepath)
    log(f"[LOAD] {sheet_type}: {filepath.name}")

    key_col = "Implementation Rate"
    for header_row in range(10):
        df = pd.read_excel(filepath, header=header_row, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
        if key_col in df.columns:
            log(f"[LOAD] 헤더 행 {header_row} 확정")
            df = df.dropna(how="all")
            return df

    raise ValueError(f"{filepath.name}: 헤더 행을 찾지 못했습니다.")


# ══════════════════════════════════════════════════════════════════════
# ▌ 2. 시트에 raw data 붙여넣기
# ══════════════════════════════════════════════════════════════════════

def write_sheet(wb: openpyxl.Workbook, sheet_name: str, df: pd.DataFrame):
    """
    워크북의 sheet_name 시트를 df 데이터로 완전히 교체.
    헤더(row 1) + 데이터(row 2~)
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    # 헤더 쓰기
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # 날짜 컬럼 인덱스 파악
    date_cols = {i for i, c in enumerate(df.columns) if 'date' in str(c).lower() or '일자' in str(c)}

    # 데이터 쓰기
    import datetime as _dt
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            # NaN → 빈칸
            if pd.isna(value):
                ws.cell(row=row_idx, column=col_idx, value=None)
            elif col_idx - 1 in date_cols:
                # 날짜 컬럼 → datetime 타입으로 변환
                try:
                    if isinstance(value, (_dt.datetime, _dt.date)):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                    else:
                        parsed = pd.to_datetime(str(value), errors='coerce')
                        if pd.notna(parsed):
                            ws.cell(row=row_idx, column=col_idx, value=parsed.to_pydatetime())
                        else:
                            ws.cell(row=row_idx, column=col_idx, value=str(value).strip())
                except:
                    ws.cell(row=row_idx, column=col_idx, value=str(value).strip())
            else:
                # 숫자 변환 시도 (Plan, Actual 등 수치 컬럼)
                try:
                    ws.cell(row=row_idx, column=col_idx, value=float(value))
                except (ValueError, TypeError):
                    ws.cell(row=row_idx, column=col_idx, value=str(value).strip())

    log(f"[WRITE] {sheet_name} 시트: {len(df)}행 작성 완료")


# ══════════════════════════════════════════════════════════════════════
# ▌ 3. LibreOffice로 수식 재계산 (선택)
# ══════════════════════════════════════════════════════════════════════

def recalc_with_libreoffice(filepath: Path) -> bool:
    """LibreOffice headless로 파일 열고 저장 → 수식 재계산"""
    lo = Path(LIBREOFFICE_PATH)
    if not lo.exists():
        log(f"[RECALC] LibreOffice 없음 → Excel에서 열면 자동 계산됩니다.")
        return False

    log(f"[RECALC] LibreOffice로 수식 재계산 중...")
    cmd = [
        str(lo), "--headless", "--invisible",
        "--convert-to", "xlsx",
        "--outdir", str(filepath.parent),
        str(filepath)
    ]
    try:
        subprocess.run(cmd, timeout=60, check=True,
                       capture_output=True)
        log(f"[RECALC] 완료")
        return True
    except Exception as e:
        log(f"[RECALC] 실패 (무시 가능): {e}")
        return False


# ══════════════════════════════════════════════════════════════════════
# ▌ 4. Rates 시트 캡처 → base64 이미지
# ══════════════════════════════════════════════════════════════════════

def capture_rates_sheet(xlsx_path: Path) -> str | None:
    """
    Excel COM으로 수식 계산 후 Rates 시트 데이터를 읽어
    HTML 테이블 문자열로 반환.
    이미지 대신 HTML 테이블 사용 → 용량 최소화, 가독성 향상.
    """
    import time
    log("[TABLE] Rates 시트 HTML 테이블 생성 시작...")

    excel = None
    wb    = None
    try:
        import win32com.client as win32
        excel = win32.DispatchEx("Excel.Application")
        wb    = excel.Workbooks.Open(str(xlsx_path.resolve()))
        ws    = wb.Worksheets("Rates")
        wb.Application.Calculate()
        time.sleep(3)

        def cell_val(r, c, is_rate=False):
            v = ws.Cells(r, c).Value
            if v is None: return ""
            if isinstance(v, float) or isinstance(v, int):
                if is_rate:
                    if isinstance(v, float) and 0 < v <= 1:
                        return f"{v*100:.1f}%"
                    if v == 0: return "-"
                    return f"{float(v):.1f}%"
                if isinstance(v, float) and v == int(v):
                    return str(int(v))
                if isinstance(v, float):
                    return f"{v:.1f}"
                return str(v)
            return str(v).strip()

        def is_red(r, c):
            try:
                interior = ws.Cells(r, c).Interior.Color
                # BGR: 빨강 계열 (낮은 달성률)
                red = interior & 0xFF
                green = (interior >> 8) & 0xFF
                blue  = (interior >> 16) & 0xFF
                return red > 150 and green < 100
            except: return False

        def rate_style(val_str):
            """달성률에 따라 배경색 반환"""
            try:
                if not val_str or val_str in ("-", ""): return ""
                v = float(str(val_str).replace('%',''))
                if v >= 100: return "background:#d1fae5;color:#065f46;font-weight:bold"
                if v >= 90:  return "background:#fef9c3;color:#854d0e;font-weight:bold"
                return "background:#fee2e2;color:#991b1b;font-weight:bold"
            except: return ""

        # ── 공통 스타일 ──────────────────────────────────────────
        TH = 'style="background:#1B2A4A;color:#fff;padding:5px 8px;font-size:11px;white-space:nowrap;text-align:center;"'
        TD = 'style="padding:4px 7px;font-size:11px;white-space:nowrap;border-bottom:1px solid #e5e7eb;"'
        TABLE_STYLE = 'style="border-collapse:collapse;width:100%;max-width:800px;font-family:Arial,sans-serif;margin-bottom:20px;"'
        SEC_HDR = 'style="background:#2E4057;color:#fff;padding:8px 12px;font-size:13px;font-weight:bold;max-width:800px;display:block;"'

        html_parts = []

        # ════════════════════════════════════════════════════════
        # 1. 상단 요약 테이블 (B2:J11)
        # ════════════════════════════════════════════════════════
        html_parts.append(f'<div {SEC_HDR[6:-1].replace("SEC_HDR","")}>📊 Completion Rate Summary</div>'.replace(
            'SEC_HDR[6:-1].replace("SEC_HDR","")','').replace('<div >','<div style="background:#2E4057;color:#fff;padding:8px 12px;font-size:13px;font-weight:bold;">'))

        html_parts.append(f'<table {TABLE_STYLE}>')
        html_parts.append(f'<tr><th {TH}>Department</th><th {TH}>LPA Plan</th><th {TH}>LPA Actual</th><th {TH}>LPA Rate</th><th {TH}></th><th {TH}>Department</th><th {TH}>5S Plan</th><th {TH}>5S Actual</th><th {TH}>5S Rate</th></tr>')

        for r in range(4, 12):
            dept_lpa = cell_val(r, 2)
            plan_lpa = cell_val(r, 3)
            act_lpa  = cell_val(r, 4)
            rate_lpa = cell_val(r, 5, is_rate=True)
            if rate_lpa == "-" and plan_lpa not in ("", "0", "-") and act_lpa in ("0", ""):
                rate_lpa = "0.0%"
            dept_5s  = cell_val(r, 7)
            plan_5s  = cell_val(r, 8)
            act_5s   = cell_val(r, 9)
            rate_5s  = cell_val(r, 10, is_rate=True)
            if rate_5s == "-" and plan_5s not in ("", "0", "-") and act_5s in ("0", ""):
                rate_5s = "0.0%"
            if not dept_lpa and not dept_5s: continue
            rs_lpa = rate_style(rate_lpa)
            rs_5s  = rate_style(rate_5s)
            html_parts.append(
                f'<tr>'
                f'<td {TD}>{dept_lpa}</td>'
                f'<td {TD} style="text-align:right;padding:4px 7px;font-size:11px;">{plan_lpa}</td>'
                f'<td {TD} style="text-align:right;padding:4px 7px;font-size:11px;">{act_lpa}</td>'
                f'<td style="padding:4px 7px;font-size:11px;text-align:center;{rs_lpa}">{rate_lpa}</td>'
                f'<td {TD}></td>'
                f'<td {TD}>{dept_5s}</td>'
                f'<td {TD} style="text-align:right;padding:4px 7px;font-size:11px;">{plan_5s}</td>'
                f'<td {TD} style="text-align:right;padding:4px 7px;font-size:11px;">{act_5s}</td>'
                f'<td style="padding:4px 7px;font-size:11px;text-align:center;{rs_5s}">{rate_5s}</td>'
                f'</tr>'
            )
        html_parts.append('</table>')

        # ════════════════════════════════════════════════════════
        # 2. 상세 테이블 — LPA (B21:U187)
        # ════════════════════════════════════════════════════════
        html_parts.append('<div style="background:#2E4057;color:#fff;padding:8px 12px;font-size:13px;font-weight:bold;">📋 LPA Detail</div>')
        html_parts.append(f'<table {TABLE_STYLE}>')
        html_parts.append(f'<tr><th {TH}>Section</th><th {TH}>Layer</th><th {TH}>Person</th><th {TH}>W/C</th><th {TH}>Plan</th><th {TH}>Actual</th><th {TH}>Rate</th></tr>')

        cur_section = ""
        for r in range(21, 188):
            b = cell_val(r, 2)   # Section / Layer
            c = cell_val(r, 3)   # Person
            d = cell_val(r, 4)   # W/C
            e = cell_val(r, 5)   # Plan
            f = cell_val(r, 6)   # Actual
            g = cell_val(r, 7, is_rate=True)   # Rate
            # Plan>0, Actual=0이면 0.0%로 표시
            if g == "-" and e not in ("", "0", "-") and f in ("0", ""):
                g = "0.0%"
            if not any([b,c,d,e,f,g]): continue
            # 섹션 헤더 감지 (Layer/Person/W/C 없고 B에 값)
            if b and not d and not e and b not in ("Layer","Subtotal","Total"):
                cur_section = b
                continue
            if b in ("Layer",): continue
            # Subtotal / Total 행
            if b in ("Subtotal","Total") or c in ("Subtotal","Total"):
                rs = rate_style(g)
                html_parts.append(
                    f'<tr style="background:#f1f5f9;font-weight:bold;">'
                    f'<td {TD}>{cur_section}</td>'
                    f'<td {TD}>{b or c}</td><td {TD}></td><td {TD}></td>'
                    f'<td {TD} style="text-align:right;padding:4px 7px;font-size:11px;">{e}</td>'
                    f'<td {TD} style="text-align:right;padding:4px 7px;font-size:11px;">{f}</td>'
                    f'<td style="padding:4px 7px;font-size:11px;text-align:center;{rs}">{g}</td>'
                    f'</tr>'
                )
                continue
            if not d and not e: continue
            rs = rate_style(g)
            html_parts.append(
                f'<tr>'
                f'<td {TD}>{cur_section}</td>'
                f'<td {TD}>{b}</td><td {TD}>{c}</td><td {TD}>{d}</td>'
                f'<td {TD} style="text-align:right;padding:4px 7px;font-size:11px;">{e}</td>'
                f'<td {TD} style="text-align:right;padding:4px 7px;font-size:11px;">{f}</td>'
                f'<td style="padding:4px 7px;font-size:11px;text-align:center;{rs}">{g}</td>'
                f'</tr>'
            )
        html_parts.append('</table>')

        # ════════════════════════════════════════════════════════
        # 3. 상세 테이블 — 5S (Y21:AR187)
        # ════════════════════════════════════════════════════════
        html_parts.append('<div style="background:#2E4057;color:#fff;padding:8px 12px;font-size:13px;font-weight:bold;">📋 5S Detail</div>')
        html_parts.append(f'<table {TABLE_STYLE}>')
        html_parts.append(f'<tr><th {TH}>Section</th><th {TH}>Layer</th><th {TH}>Person</th><th {TH}>W/C</th><th {TH}>Plan</th><th {TH}>Actual</th><th {TH}>Rate</th></tr>')

        cur_section = ""
        for r in range(21, 188):
            b = cell_val(r, 25)  # Y — Layer
            c = cell_val(r, 26)  # Z — Person
            d = cell_val(r, 27)  # AA — W/C
            e = cell_val(r, 28)  # AB — Plan
            f = cell_val(r, 29)  # AC — Actual
            g = cell_val(r, 30, is_rate=True)  # AD — Rate
            # Plan>0, Actual=0이면 0.0%로 표시
            if g == "-" and e not in ("", "0", "-") and f in ("0", ""):
                g = "0.0%"
            if not any([b,c,d,e,f,g]): continue
            if b and not d and not e and b not in ("Layer","Subtotal","Total"):
                cur_section = b
                continue
            if b in ("Layer",): continue
            if b in ("Subtotal","Total") or c in ("Subtotal","Total"):
                rs = rate_style(g)
                html_parts.append(
                    f'<tr style="background:#f1f5f9;font-weight:bold;">'
                    f'<td {TD}>{cur_section}</td>'
                    f'<td {TD}>{b or c}</td><td {TD}></td><td {TD}></td>'
                    f'<td {TD} style="text-align:right;padding:4px 7px;font-size:11px;">{e}</td>'
                    f'<td {TD} style="text-align:right;padding:4px 7px;font-size:11px;">{f}</td>'
                    f'<td style="padding:4px 7px;font-size:11px;text-align:center;{rs}">{g}</td>'
                    f'</tr>'
                )
                continue
            if not d and not e: continue
            rs = rate_style(g)
            html_parts.append(
                f'<tr>'
                f'<td {TD}>{cur_section}</td>'
                f'<td {TD}>{b}</td><td {TD}>{c}</td><td {TD}>{d}</td>'
                f'<td {TD} style="text-align:right;padding:4px 7px;font-size:11px;">{e}</td>'
                f'<td {TD} style="text-align:right;padding:4px 7px;font-size:11px;">{f}</td>'
                f'<td style="padding:4px 7px;font-size:11px;text-align:center;{rs}">{g}</td>'
                f'</tr>'
            )
        html_parts.append('</table>')

        html = '\n'.join(html_parts)
        log(f"[TABLE] HTML 테이블 생성 완료 ({len(html):,} chars)")
        return html

    except Exception as e:
        log(f"[TABLE] 오류: {e}")
        import traceback; traceback.print_exc()
        return None
    finally:
        try:
            if wb:    wb.Close(False)
            if excel: excel.Quit()
        except: pass


# ══════════════════════════════════════════════════════════════════════
# ▌ 5. 메인 업데이트 함수
# ══════════════════════════════════════════════════════════════════════

def update_report(lpa_file: Path, s5_file: Path, from_d=None, to_d=None) -> tuple[Path, str | None]:
    """
    템플릿을 복사하고 LPA/5S 시트를 raw data로 교체.
    Rates 시트의 SUMIFS 수식은 그대로 유지 → 자동 재계산.

    Returns: (저장된 xlsx 경로, Rates 시트 base64 이미지 or None)
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # ── 템플릿 복사 ──────────────────────────────────────────────────
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = OUTPUT_DIR / f"Plant_Mobile_Audit_Report_{ts}.xlsx"
    shutil.copy2(str(TEMPLATE_PATH), str(out_path))
    log(f"[TEMPLATE] 복사: {out_path.name}")

    # ── raw data 로드 ─────────────────────────────────────────────────
    df_lpa = load_raw(lpa_file, "LPA")
    df_5s  = load_raw(s5_file,  "5S")

    # ── 날짜 필터링 (from_d ~ to_d) ──────────────────────────────────
    if from_d and to_d:
        import pandas as pd
        for df, date_col in [(df_lpa, "Plan Date"), (df_5s, "Plan Date")]:
            if date_col in df.columns:
                df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
        # LPA 필터
        if "Plan Date" in df_lpa.columns:
            mask = (df_lpa["Plan Date"] >= pd.Timestamp(from_d)) &                    (df_lpa["Plan Date"] <= pd.Timestamp(to_d))
            df_lpa = df_lpa[mask].copy()
            log(f"[FILTER] LPA 날짜 필터: {len(df_lpa)}행")
        # 5S 필터
        if "Plan Date" in df_5s.columns:
            mask = (df_5s["Plan Date"] >= pd.Timestamp(from_d)) &                    (df_5s["Plan Date"] <= pd.Timestamp(to_d))
            df_5s = df_5s[mask].copy()
            log(f"[FILTER] 5S 날짜 필터: {len(df_5s)}행")

    # ── 워크북 열기 ───────────────────────────────────────────────────
    wb = load_workbook(str(out_path))

    # 시트 순서 기억 (나중에 복원)
    original_order = wb.sheetnames[:]

    # ── LPA / 5S 시트 교체 ───────────────────────────────────────────
    write_sheet(wb, "LPA", df_lpa)
    write_sheet(wb, "5S",  df_5s)

    # ── 불필요한 시트 삭제 (Chart, TPM, TBM, Process Check) ─────────
    sheets_to_delete = ["Chart", "TPM", "TBM", "Process Check"]
    for sheet_name in sheets_to_delete:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            log(f"[DELETE] '{sheet_name}' 시트 삭제")

    # ── 시트 순서 정렬: Rates → LPA → 5S ────────────────────────────
    desired_order = ["Rates", "LPA", "5S"]
    for i, sheet_name in enumerate(desired_order):
        if sheet_name in wb.sheetnames:
            wb.move_sheet(sheet_name, offset=-(len(wb.sheetnames) - i))

    # ── 저장 ──────────────────────────────────────────────────────────
    wb.save(str(out_path))
    log(f"[SAVE] 저장 완료: {out_path.name}")

    # ── LibreOffice 수식 재계산 ───────────────────────────────────────
    recalc_with_libreoffice(out_path)

    # ── Rates 시트 이미지 캡처 ────────────────────────────────────────
    img_b64 = capture_rates_sheet(out_path)

    return out_path, img_b64


# ══════════════════════════════════════════════════════════════════════
# ▌ CLI
# ══════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="LPA/5S raw data를 Plant_Mobile_Audit_Report.xlsx에 자동 업데이트"
    )
    parser.add_argument("--lpa", type=str, required=True, help="LPA raw data 파일 경로")
    parser.add_argument("--5s",  type=str, required=True, dest="s5",
                        help="5S raw data 파일 경로")
    args = parser.parse_args()

    out_path, img_b64 = update_report(
        lpa_file = Path(args.lpa),
        s5_file  = Path(args.s5),
    )

    log(f"\n✅ 완료: {out_path}")
    if img_b64:
        log("✅ Rates 시트 이미지 캡처 성공")
    else:
        log("ℹ️  이미지 캡처 없음 — Excel에서 파일 확인하세요.")


if __name__ == "__main__":
    main()
